from io import BytesIO

from openpyxl import load_workbook


def file_iter(file: bytes):
    res = []
    wb = load_workbook(BytesIO(file))
    for ws in wb.worksheets:
        title, *rows = list(ws.iter_rows(values_only=True))
        for row in rows:
            res.append(dict(zip(title, row)))
    return res
